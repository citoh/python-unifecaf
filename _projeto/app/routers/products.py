# Importações principais do FastAPI
from fastapi import APIRouter, Request, Depends, Form, HTTPException, UploadFile, File
from fastapi.responses import RedirectResponse, JSONResponse
from fastapi.templating import Jinja2Templates
from starlette import status

# ORM e utilitários de banco de dados
from sqlalchemy.orm import Session
from sqlalchemy import select
from pathlib import Path

# Manipulação de arquivos
from io import BytesIO
from openpyxl import load_workbook

# Importações do projeto
from app.database import get_db, engine, Base
from app.models import Product, User, Cart, CartItem  # Importa todos os modelos para criar as tabelas
from app.dependencies import require_admin, get_current_user_optional, get_current_user

from app.helpers import format_brl_price, format_brl_date, parse_brl_price

# Cria o roteador de produtos
router = APIRouter()

# Cria as tabelas do banco na primeira importação do módulo
Base.metadata.create_all(bind=engine)

# Configuração de diretório de templates
TEMPLATES_DIR = Path(__file__).resolve().parent.parent / "templates"
templates = Jinja2Templates(directory=str(TEMPLATES_DIR))
templates.env.filters["brl_price"] = format_brl_price
templates.env.filters["brl_date"] = format_brl_date


# ------------------------------------------------------------------
# Redireciona a raiz "/" - removido, agora está no main.py
# ------------------------------------------------------------------


# ------------------------------------------------------------------
# LISTAR PRODUTOS (GET)
# Exibe a lista completa de produtos cadastrados
# ------------------------------------------------------------------
@router.get("/products")
def list_products(request: Request, db: Session = Depends(get_db), current_user = Depends(get_current_user)):
    """Lista produtos. Admin vê gerenciamento completo, user vê lista para compra."""
    # Busca todos os produtos, ordenando do mais recente para o mais antigo
    products = db.scalars(select(Product).order_by(Product.id.desc())).all()
    # Renderiza a página de listagem
    return templates.TemplateResponse(
        "products/index.html",
        {"request": request, "products": products, "current_user": current_user}
    )


# ------------------------------------------------------------------
# FORMULÁRIO DE NOVO PRODUTO (GET)
# Exibe o formulário vazio para cadastro de um novo produto
# ------------------------------------------------------------------
@router.get("/products/new")
def new_product_form(request: Request, current_user = Depends(require_admin)):
    return templates.TemplateResponse(
        "products/new.html",
        {
            "request": request,
            "action": "/products",             # rota de envio do formulário
            "method_override": "POST",         # método HTTP utilizado
            "product": None,                   # sem produto (novo cadastro)
            "current_user": current_user
        }
    )


# ------------------------------------------------------------------
# FORMULÁRIO DE EDIÇÃO DE PRODUTO (GET)
# Exibe o formulário preenchido com os dados de um produto existente
# ------------------------------------------------------------------
@router.get("/products/{product_id}/edit")
def edit_product_form(product_id: int, request: Request, db: Session = Depends(get_db), current_user = Depends(require_admin)):
    # Busca o produto pelo ID
    product = db.get(Product, product_id)
    if not product:
        raise HTTPException(status_code=404, detail="Produto não encontrado")

    # Renderiza o formulário com dados preenchidos
    return templates.TemplateResponse(
        "products/edit.html",
        {
            "request": request,
            "action": f"/products/{product.id}",  # rota para envio da atualização
            "method_override": "PUT",             # método de atualização
            "product": product,
            "current_user": current_user
        }
    )


# ------------------------------------------------------------------
# CRIAÇÃO DE PRODUTO (POST)
# Recebe os dados do formulário e insere um novo registro no banco
# ------------------------------------------------------------------
@router.post("/products")
def create_product(
    name: str = Form(...),
    sku: str = Form(...),
    price: float = Form(...),
    db: Session = Depends(get_db),
    current_user = Depends(require_admin),
):
    # Validação simples de campos obrigatórios
    if not name.strip() or not sku.strip():
        raise HTTPException(status_code=400, detail="Nome e SKU são obrigatórios")

    # Verifica se já existe um produto com o mesmo SKU
    if db.scalar(select(Product).where(Product.sku == sku)):
        raise HTTPException(status_code=400, detail="SKU já cadastrado")

    # Cria o objeto e salva no banco
    db.add(Product(name=name.strip(), sku=sku.strip(), price=price))
    db.commit()

    # Redireciona para a listagem de produtos
    return RedirectResponse(url="/products", status_code=status.HTTP_303_SEE_OTHER)


# ------------------------------------------------------------------
# ATUALIZAÇÃO DE PRODUTO (PUT)
# Atualiza os dados de um produto existente
# ------------------------------------------------------------------
@router.put("/products/{product_id}")
def update_product(
    product_id: int,
    name: str = Form(...),
    sku: str = Form(...),
    price: float = Form(...),
    db: Session = Depends(get_db),
    current_user = Depends(require_admin),
):
    # Busca o produto pelo ID
    product = db.get(Product, product_id)
    if not product:
        raise HTTPException(status_code=404, detail="Produto não encontrado")

    # Verifica se o SKU já pertence a outro produto
    if db.scalar(select(Product).where(Product.sku == sku, Product.id != product_id)):
        raise HTTPException(status_code=400, detail="SKU já cadastrado para outro produto")

    # Atualiza os campos
    product.name = name.strip()
    product.sku = sku.strip()
    product.price = price

    # Confirma no banco
    db.add(product)
    db.commit()

    # Retorna sucesso sem conteúdo (204)
    return JSONResponse(status_code=status.HTTP_204_NO_CONTENT, content=None)


# ------------------------------------------------------------------
# EXCLUSÃO DE PRODUTO (DELETE)
# Remove um produto do banco de dados
# ------------------------------------------------------------------
@router.delete("/products/{product_id}")
def delete_product(product_id: int, db: Session = Depends(get_db), current_user = Depends(require_admin)):
    # Busca o produto pelo ID
    product = db.get(Product, product_id)
    if not product:
        raise HTTPException(status_code=404, detail="Produto não encontrado")

    # Remove e confirma
    db.delete(product)
    db.commit()

    # Retorna sucesso sem conteúdo (204)
    return JSONResponse(status_code=status.HTTP_204_NO_CONTENT, content=None)


# ------------------------------------------------------------------
# IMPORTAR DADOS DO EXCEL
# Importar produtos a partir de uma tabela criada no Excel
# ------------------------------------------------------------------

# Formulário de Importação    
@router.get("/products/import")
def import_products_form(request: Request, current_user = Depends(require_admin)):
    return templates.TemplateResponse(
        "products/import.html",
        {"request": request, "report": None, "current_user": current_user}
    )


# Importando produtos de um arquivo em excel
@router.post("/products/import")
async def import_products(
    request: Request,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user = Depends(require_admin),
):
    """
    Recebe um arquivo .xlsx com colunas obrigatórias:
    Nome, Sku e Preço.
    Valida os dados e importa apenas produtos válidos.
    """

    # Verifica tipo de arquivo
    allowed = {
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "application/vnd.ms-excel",
        "application/octet-stream",
    }
    if file.content_type not in allowed:
        raise HTTPException(status_code=400, detail="Arquivo inválido. Envie um .xlsx")

    # Lê o conteúdo em memória
    content = await file.read()
    try:
        wb = load_workbook(BytesIO(content), data_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="Não foi possível ler o arquivo .xlsx")

    ws = wb.active
    if ws.max_row < 2:
        return templates.TemplateResponse(
            "products/import.html",
            {"request": request, "report": {"imported": 0, "skipped": 0, "errors": [{"row": 1, "error": "Planilha vazia"}]}, "current_user": current_user}
        )

    # -------------------------------------------------------------
    # Validação estrita de cabeçalhos
    # Espera-se: Nome | Sku | Preço (nessa ordem)
    # -------------------------------------------------------------
    headers = [str(ws.cell(row=1, column=c).value or "").strip() for c in range(1, ws.max_column + 1)]
    expected_headers = ["Nome", "Sku", "Preço"]

    if headers[:3] != expected_headers:
        raise HTTPException(
            status_code=400,
            detail=f"Cabeçalhos inválidos. Esperado: {', '.join(expected_headers)}"
        )

    # Mapear colunas exatas
    header_map = {"name": 1, "sku": 2, "price": 3}

    # -------------------------------------------------------------
    # Preparar validações e variáveis
    # -------------------------------------------------------------
    existing_skus = set(db.scalars(select(Product.sku)).all())
    seen_skus_in_file = set()

    imported = 0
    skipped = 0
    errors = []
    new_products = []

    # -------------------------------------------------------------
    # Iterar linhas de dados (a partir da linha 2)
    # -------------------------------------------------------------
    for row in range(2, ws.max_row + 1):
        try:
            raw_name = ws.cell(row=row, column=header_map["name"]).value
            raw_sku = ws.cell(row=row, column=header_map["sku"]).value
            raw_price = ws.cell(row=row, column=header_map["price"]).value

            name = (str(raw_name).strip() if raw_name else "")
            sku = (str(raw_sku).strip() if raw_sku else "")

            if not name:
                raise ValueError("Nome vazio")
            if not sku:
                raise ValueError("SKU vazio")

            price = parse_brl_price(raw_price)

            if sku in existing_skus:
                raise ValueError("SKU já cadastrado no banco")
            if sku in seen_skus_in_file:
                raise ValueError("SKU duplicado no arquivo")

            new_products.append(Product(name=name, sku=sku, price=price))
            seen_skus_in_file.add(sku)
            imported += 1

        except Exception as e:
            skipped += 1
            errors.append({"row": row, "error": str(e)})

    # Inserir produtos válidos
    if new_products:
        db.add_all(new_products)
        db.commit()

    # Retornar relatório
    report = {"imported": imported, "skipped": skipped, "errors": errors}
    return templates.TemplateResponse(
        "products/import.html",
        {"request": request, "report": report, "current_user": current_user}
    )
    
# ------------------------------------------------------------------
# DETALHE DE PRODUTO (GET)
# Exibe os dados detalhados de um produto específico
# ------------------------------------------------------------------
@router.get("/products/{product_id}")
def get_product(product_id: int, request: Request, db: Session = Depends(get_db), current_user = Depends(get_current_user)):
    # Busca o produto pelo ID
    product = db.get(Product, product_id)
    if not product:
        raise HTTPException(status_code=404, detail="Produto não encontrado")

    # Renderiza a página de detalhes
    return templates.TemplateResponse(
        "products/show.html",
        {"request": request, "product": product, "current_user": current_user}
    )